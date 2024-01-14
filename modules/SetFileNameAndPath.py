from datetime import date
import datetime
import openpyxl
import pandas as pd

class SetFileNameAndPath:
    def __init__(self):
        today_date = date.today()
        self.today_date_formatted = '{}.{}.{}'.format(today_date.day, today_date.month, today_date.year)
        self.file_name = 'products_inventory-'+self.today_date_formatted+'.xlsx' 
        self.file_name_and_path = 'ParametersFiles\\'+self.file_name 

    def set_file_name_to_work_with(self):
        try:
            self.workbook = openpyxl.load_workbook('ParametersFiles\DateParameter.xlsx')
            self.worksheet = self.workbook['Plan1']
            self.number_of_rows = int(self.worksheet.max_row) - 1
            if self.number_of_rows > 0:
                for row in self.worksheet.iter_rows(min_row=2):
                    self.date_value = row[0].value
                    if self.date_value is not None:
                        if isinstance(self.date_value, datetime.datetime):
                            self.date_value = self.date_value.date()
                        self.date_value_formatted = '{}.{}.{}'.format(self.date_value.day, self.date_value.month, self.date_value.year)
                        self.file_name = 'products_inventory-'+self.date_value_formatted+'.xlsx'
                        self.file_name_and_path = 'ParametersFiles\\'+self.file_name
                      
            else:
                self.file_name = 'products_inventory-'+self.today_date_formatted+'.xlsx'
                self.file_name_and_path = 'ParametersFiles\\'+self.file_name 
                       
            self.data_frame = pd.read_excel(self.file_name_and_path, sheet_name='vendas')        
           
            return self.file_name_and_path, self.data_frame
        
        except Exception as e:
            self.set_log.set_log_error(str(e))
