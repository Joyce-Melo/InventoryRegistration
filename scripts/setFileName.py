from datetime import date
import datetime

import openpyxl

import pandas as pd 


def set_file_name():
    today = date.today() #YYYY/MM/DD
    print('Data antes de ser formatada', today) 
    today_formatted = '{}.{}.{}'.format(today.day, today.month, today.year)
    print('Data após formatada', today_formatted)
    file_name = 'products_inventory-'+today_formatted+'.xlsx'
    print('Nome do arquivo', file_name)
    file_name_and_path = 'ParametersFiles\\'+file_name
        #Ler a data qeu está no arquivo DateParameter
    workbook = openpyxl.load_workbook('ParametersFiles\DateParameter.xlsx')
    worksheet = workbook['Plan1']
    number_of_rows = int(worksheet.max_row) - 1
        # print('O número de linhas (excluindo o cabeçalho) do arquivo é: ',number_of_rows) #imprimindo o número de linhas menos o cabeçalho
    if number_of_rows > 0 :
        print('entrou na condição')
        for row in worksheet.iter_rows(min_row=2):#pulando o cabeçalho
                date_value = row[0].value # Para mais de uma data aqui provavelmente você terá que ir incrmentando 1 
                if isinstance(date_value, datetime.datetime):
                    date_value = date_value.date()  # Converter para um objeto date
                print(date_value)

        date_value_formatted = '{}.{}.{}'.format(date_value.day, date_value.month, date_value.year)
        print('A data do arquivo após formatada é: ', date_value_formatted)

        file_name = 'products_inventory-'+date_value_formatted+'.xlsx'
        print('O nome do arquivo com a data do arquivo de parametros é: ', file_name)
        file_name_and_path = 'ParametersFiles\\'+file_name
    else:
        file_name = 'products_inventory-'+today_formatted+'.xlsx'
        print('O nome do arquivo com a data do arquivo de parametros é: ', file_name)
        file_name_and_path = 'ParametersFiles\\'+file_name

    return file_name_and_path

set_file_name()